import os
import xlwt
from openpyxl import load_workbook
import openpyxl
import os


def calculate_tpfptnfn_kframe(tru_txt_path, pre_txt_path, file_pre, file_num, unit_size, a, b, 
        conf_thre=0.1, iou_thre=0.1):
    """
    use k-frame method to compute TP, TN, FP, FN
    :param  file_pre: ep."place1_whitecane01.mp4_5.txt" pre is "place1_whitecane01.mp4"
            unit_size: how many flame compose an unit, = k.
            file_num: how many flame we have.
            a: the error modulus to judge true, false. from 1/k to 1.
            b: the error modulus to judge positive, negative. from 1/k to 1.
            conf_thre: confidence thread to abandon unreliable flame.
            iou_thre: iou thread to delete wrong results.
    :return f-score
    """

    # default some varities.
    tp = 0
    tn = 0
    fp = 0
    fn = 0
    unit_num = file_num // unit_size
    folder_pre_name = pre_txt_path
    folder_tru_name = tru_txt_path
    # start = os.listdir(tru_txt_path)[0][-5]


    # for each unit.
    # use flame_truth_num and flame_pre_num to calculate the tp, tn, fp, fn.
    for i in range(unit_num+1):

        # count the number of flame of predition and truth.
        flame_truth_num = 0
        flame_pre_num = 0
        
        # for each flame.
        # iterate an unit, and classify which status(tp, tn, fp, fn) it is.
        for j in range(1, unit_size + 1):
            
            # find the label txt. 
            if (j + i * unit_size) > 99:
                txtname = file_pre + '_{}.txt'.format(j + i * unit_size)
            elif 10 <= (j + i * unit_size) <= 99:
                txtname = file_pre + '_0{}.txt'.format(j + i * unit_size)
            else:
                txtname = file_pre + '_00{}.txt'.format(j + i * unit_size)
            pre_txt_path = os.path.join(folder_pre_name, txtname)
            tru_txt_path = os.path.join(folder_tru_name, txtname)


            # this flame actuallty has whitecane or not.
            if os.path.exists(tru_txt_path):
                flame_truth_num += 1
            # pre label必须存在且不为空，才能说明yolo检测出目标了。
            if os.path.exists(pre_txt_path):
                with open(pre_txt_path, '+r') as f:
                    t = f.read()
                # yolo检测出错误的目标，实际没有。
                if t and not os.path.exists(tru_txt_path):
                    flame_pre_num += 1  # 变成FP
                # yolo检测出正确目标，且实际有。
                elif t and os.path.exists(tru_txt_path) and get_iou(tru_txt_path, pre_txt_path, conf_thre) >= iou_thre:
                    flame_pre_num += 1  # 变成TP必须要正确检测
  
        # calculate the tp, tn, fp, fn.
        if flame_truth_num >= (unit_size * a) and flame_pre_num >= (unit_size * b):
            tp += 1
        elif flame_truth_num >= (unit_size * a) and flame_pre_num < (unit_size * b):
            fn += 1
        elif flame_truth_num < (unit_size * a) and flame_pre_num >= (unit_size * b):
            fp += 1       
        elif flame_truth_num < (unit_size * a) and flame_pre_num < (unit_size * b):
            tn += 1
        
    return (tp,fp,tn,fn)

def calculate_fscore(tp, tn, fp, fn):
    """
    compute f-score
    """
    precision = tp / (tp + fp) if (tp + fp) != 0 else 0
    recall = tp / (tp + fn) if (tp + fn) != 0 else 0
    return (2 * precision * recall) / (precision + recall) if (precision + recall) != 0 else 0

def compute_iou(b1, b2):
    """
    compute IoU
    :param b1: [y0, x0, y1, x1], which reflect box's 
                bottom line, left line, top line, right line.
           b2: [y0, x0, y1, x1]
    :return scale value of IoU
    """
    # compute area of each 
    s_b1 = (b1[2] - b1[0]) * (b1[3] - b1[1])
    s_b2 = (b2[2] - b2[0]) * (b2[3] - b2[1])

    # compute the sum_area
    sum_area = s_b1 + s_b2

    # compute intersect area
    bottom_line = max(b1[0], b2[0])
    left_line = max(b1[1], b2[1])
    top_line = min(b1[2], b2[2])
    right_line = min(b1[3], b2[3])

    # judge if there is an intersect
    if bottom_line >= top_line or left_line >= right_line:
        return 0
    else:
        intersect_area = (top_line - bottom_line) * (right_line - left_line)
        return (intersect_area/(sum_area - intersect_area))

def get_tru_boxes(file):
    """
    extract box information from txt
    :param file: "xxx.txt"
    :return [[y0, x0, y1, x1],]
            which reflect boxes's bottom line, left line, top line, right line.
    """
    # read txt and split different boxes
    txt = open(file, 'r').read().split('\n')
    
    # create the boxes list
    boxes = []

    # add boxes information into list
    # pre_txt: [n, x, y, w, h, c] to [bottom line, left line, top line, right line, c]
    # tru_txt: [n, x, y, w, h] to [bottom line, left line, top line, right line]
    for box in txt:
        if box:
            box = box.split(' ')
            bottom_line = float(box[2]) - float(box[4])/2
            left_line = float(box[1]) - float(box[3])/2
            top_line = float(box[2]) + float(box[4])/2
            right_line = float(box[1]) + float(box[3])/2
            
            boxes.append([bottom_line, left_line, top_line, right_line])

    return boxes

def get_pre_boxes(file):
    """
    extract box information from txt
    :file: file path
    :param file: "xxx.txt"
    :return [[y0, x0, y1, x1, c],]
            which reflect boxes's bottom line, left line, top line, right line.
    """
    # read txt and split different boxes
    txt = open(file, 'r').read().split('\n')
    
    # create the boxes list
    boxes = []

    # add boxes information into list
    # pre_txt: [n, x, y, w, h, c] to [bottom line, left line, top line, right line, c]
    # tru_txt: [n, x, y, w, h] to [bottom line, left line, top line, right line]
    for box in txt:
        if box:
            box = box.split(' ')
            bottom_line = float(box[2]) - float(box[4])/2
            left_line = float(box[1]) - float(box[3])/2
            top_line = float(box[2]) + float(box[4])/2
            right_line = float(box[1]) + float(box[3])/2
            
            boxes.append([bottom_line, left_line, top_line, right_line, float(box[-1])])

    return boxes

def get_iou(tru_txt, pre_txt, conf_thre):
    """
    compute iou from two result files.
    :param tru_txt, pre_txt: "xxx.txt"
    :return max iou
    """
    t_boxes = get_tru_boxes(tru_txt)
    p_boxes = get_pre_boxes(pre_txt)
    iou = 0

    # iterate every box in two file and compute all their iou
    for t_box in t_boxes:
        for p_box in p_boxes:
            if p_box[-1] >= conf_thre:
                new_iou = compute_iou(t_box, p_box)
                iou = max(iou, new_iou)

    return iou

def default_excel(dayornight, k, conf_thre, iou_thre):
    """
    write the default form of k-frame excel.
    :param 
    :return excel_path
    """
    # create 'result' filter
    if not os.path.exists('result'):
        os.mkdir('result')
    
     # default format 
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    
    row1 = [f'β={i}' for i in range(1, k+1)]
    col1 = [f'α={i}' for i in range(1, k+1)]

    # write default format
    for i in range(len(row1)):
        worksheet.cell(1, i+2).value = row1[i]
    for i in range(len(col1)):
        worksheet.cell(i+2, 1).value = col1[i]
    
    # save excel file
    name = f'({dayornight}k{k})k-frame_c{conf_thre}i{iou_thre}.xlsx'
    workbook.save(fr'result/{name}')
    print('已新建',name)
    
    excel_path = os.path.join('result', name)
    return excel_path

def write_excel_by_row(excel, res):
    """
    write the results of k-frame to excel file.
    :param results: [[k, a, b, f-score],]
    :return none
    """

    workbook = load_workbook(excel)
    worksheet = workbook.worksheets[0]

    # write k-frame results
    for result in res:
        worksheet.cell(result[1]+1, result[2]+1).value = result[3]
    
    # save excel file
    workbook.save(excel)

def write_average_fscore(excel):
    """
    计算所有场景的加权平均f值。
    :param 
    :return 
    """
        
    # default format 
    workbook = load_workbook(excel)
    worksheet = workbook.worksheets[0]
    
    # 计算加权平均f值
    f = 0
    num = 0
    for row in range(2, worksheet.max_row+1):
        f = f + float(worksheet.cell(row, 2).value) * float(worksheet.cell(row, 5).value)
        num += float(worksheet.cell(row, 2).value)
    f = round(f/num, 2)
    # write f
    worksheet.cell(2, 7).value = f

    # save excel file
    workbook.save(excel)
    print(excel,f'完成计算。平均f值为{f}')
    print('-----------------------end----------------------------')


if __name__ == '__main__':
    excel = default_excel('day', 4, 0.1, 0.1)
    write_excel_by_row(excel, [[4,2,3,0.6]])